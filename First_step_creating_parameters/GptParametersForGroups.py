import openpyxl
from First_step_creating_parameters.Excel_can_crawl_lists import ExcelSheetManager
from First_step_creating_parameters.First_GPT_Agent import GptService


class GptCombinedProcessor:
    """
    Класс для обработки глобального словаря параметров и частот,
    объединения его с данными из первых 20 строк каждого листа,
    передачи в GPT сервис и записи результата в столбец J Excel-файла.
    """

    def __init__(self, excel_filename, system_prompt_file_name, basic_prompt_file_name, gpt_service=None):
        """
        Инициализация объекта класса.

        :param excel_filename: Имя файла Excel.
        :param system_prompt_file_name: Название файла системного промпта.
        :param basic_prompt_file_name: Название файла базового промпта.
        :param gpt_service: Экземпляр класса GptService для взаимодействия с GPT. Если None, создаётся новый.
        """
        self.excel_manager = ExcelSheetManager(excel_filename)
        self.excel_filename = excel_filename
        self.system_prompt_file_name = system_prompt_file_name
        self.basic_prompt_file_name = basic_prompt_file_name
        self.gpt_service = gpt_service if gpt_service else GptService()

    def process_combined_data(self, data_dict, data_strings):
        """
        Обрабатывает глобальный словарь параметров и частот,
        объединяет его с данными строк, передаёт в GPT и записывает результат в столбец J каждого листа.

        :param data_dict: Глобальный словарь с параметрами и их частотами.
        :param data_strings: Словарь, где ключ — название листа, а значение — строка с данными первых 20 строк листа.
        """
        try:
            # Загружаем Excel-файл с возможностью записи
            workbook = openpyxl.load_workbook(self.excel_filename)

            # Преобразуем глобальный словарь в строку
            data_dict_string = self.convert_dict_to_string(data_dict)
            print(f"Преобразованный глобальный словарь в строку: {data_dict_string}")

            # Получаем список всех group_id из Excel (предполагается, что они соответствуют названиям листов)
            group_ids = self.excel_manager.get_group_ids_from_sheet()
            print(f"Получено {len(group_ids)} group_id из Excel.")

            for group_id in group_ids:
                sheet_name = str(group_id)
                if sheet_name not in workbook.sheetnames:
                    print(f"Лист '{sheet_name}' не найден в Excel-файле. Пропускаем.")
                    continue

                sheet = workbook[sheet_name]
                print(f"Обрабатываем лист: {sheet_name}")

                # Добавляем заголовок в столбец J, если его нет
                header_J = sheet.cell(row=1, column=10)  # Столбец J
                if header_J.value != "GPT Анализ":
                    header_J.value = "GPT Анализ"
                    print(f"Заголовок столбца J добавлен на листе '{sheet_name}'.")

                # Получаем строку данных из первых 20 строк листа
                data_string = data_strings.get(sheet_name, "")
                if not data_string:
                    print(f"Данные из первых 20 строк для листа '{sheet_name}' отсутствуют. Используем только data_dict.")
                    combined_input = data_dict_string
                else:
                    # Объединяем data_dict и data_string
                    combined_input = f"{data_dict_string}\nДанные первых 20 строк листа '{sheet_name}':\n{data_string}"
                    print(f"Объединенные данные для листа '{sheet_name}': {combined_input}")

                # Передаём объединённую строку в GPT
                try:
                    result = self.gpt_service.gpt(
                        input_data=combined_input,
                        system_prompt_file_name=self.system_prompt_file_name,
                        basic_prompt_file_name=self.basic_prompt_file_name
                    )
                except Exception as gpt_error:
                    print(f"Ошибка при обращении к GPT для листа '{sheet_name}': {gpt_error}")
                    parsed_result = "Ошибка GPT"
                else:
                    # Получаем parsed_result, если он есть
                    parsed_result = result.get('parsed_result', "Неизвестно")
                    if not parsed_result:
                        parsed_result = "Неизвестно"

                # Записываем parsed_result в ячейку J2
                sheet.cell(row=2, column=10, value=parsed_result)

                # Сохраняем файл после вставки значения
                try:
                    workbook.save(self.excel_filename)
                    print(f"Записан результат GPT на лист '{sheet_name}' в ячейку J2. Файл сохранён.")
                except Exception as save_error:
                    print(f"Ошибка при сохранении файла для листа '{sheet_name}': {save_error}")
                    continue

                # Выводим результат в консоль (опционально)
                print(f"Лист: {sheet_name}, GPT Result: {parsed_result}, Token Usage: {result.get('token_usage', 'N/A')}")

            print("Все словари успешно обработаны и сохранены.")

        except Exception as e:
            print("Произошла ошибка при обработке словарей:", e)

    @staticmethod
    def convert_dict_to_string(param_freq_dict):
        """
        Преобразует глобальный словарь параметров и их частот в строку.

        :param param_freq_dict: Глобальный словарь с параметрами и их частотами.
        :return: Строка, сформированная из ключей и значений словаря.
        """
        return ", ".join([f"{key}: {value}" for key, value in param_freq_dict.items()])


# Пример использования
if __name__ == "__main__":
    excel_filename = 'main_data_with_analysis.xlsx'
    system_prompt_file_name = "group_parameters_system_prompt"
    basic_prompt_file_name = "group_parameters_basic_prompt"

    # Создаем экземпляр GptService
    gpt_service = GptService()

    # Создаем экземпляр GptCombinedProcessor
    processor = GptCombinedProcessor(
        excel_filename=excel_filename,
        system_prompt_file_name=system_prompt_file_name,
        basic_prompt_file_name=basic_prompt_file_name,
        gpt_service=gpt_service
    )

    # Пример глобального словаря для обработки
    data_dict = {
        "volume": 111,
        "color": 13,
        "completeness": 168,
        "category": 469,
        "signs_of_difference": 3
        # Добавьте другие параметры по необходимости
    }

    # Пример словаря с данными первых 20 строк для каждого листа
    # Формат: {"14": "ID\tНаименование\tМаркировка\tРегламенты (ГОСТ/ТУ)\tПараметры\tOKPD2_NAME\tГОСТ Название\n8559230749\t...", ...}
    data_strings = {
        "14": """ID\tНаименование\tМаркировка\tРегламенты (ГОСТ/ТУ)\tПараметры\tOKPD2_NAME\tГОСТ Название
8559230749\tСОРОЧКА МУЖСКАЯ ФОРМЕННАЯ ПОВСЕДНЕВНАЯ С КОРОТКИМИ РУКАВАМИ ДЛЯ НАЧАЛЬНИКОВ ПОЕЗДОВ, ПРОВОДНИКОВ И КАССИРОВ\tТИП Г\tТУ 14.14.21-010-94154560-2021\t124-126-188 БЕЛАЯ В КОРИЧНЕВО-КРАСНУЮ ПОЛОСКУ (5-АЯ ПОЛНОТНАЯ ГРУППА), СРЕДНЯЯ КАТЕГОРИЯ ДОЛЖНОСТИ, 2 ГАЛУНА 3 ЗВЕЗДЫ\tХалаты, фартуки, жилеты и сорочки мужские производственные и профессиональные\t""",
        # Добавьте другие листы и их данные по аналогии
    }

    # Запускаем процесс обработки данных
    processor.process_combined_data(data_dict, data_strings)
