import openpyxl
import pyodbc
from openpyxl.utils import get_column_letter
from collections import defaultdict
from First_step_creating_parameters.Excel_can_crawl_lists import ExcelSheetManager


class ExcelGatherDataOptimized:
    """
    Оптимизированный класс для работы с базой данных Access, объединения таблиц и записи данных в Excel.
    """

    def __init__(self, db_path, excel_filename):
        """
        Инициализация объекта класса.

        :param db_path: Путь к базе данных Access.
        :param excel_filename: Имя файла Excel.
        """
        self.db_path = db_path
        self.excel_manager = ExcelSheetManager(excel_filename)
        self.conn = self.connect_to_db()
        self.data_mapping = defaultdict(lambda: [""] * 10)  # Инициализируем словарь с пустыми значениями

    def connect_to_db(self):
        """
        Устанавливает соединение с базой данных Access.

        :return: Объект соединения.
        """
        conn_str = (
            r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
            r'DBQ=' + self.db_path + ';'
        )
        try:
            conn = pyodbc.connect(conn_str)
            print("Соединение с базой данных успешно установлено.")
            return conn
        except pyodbc.Error as e:
            print("Ошибка соединения с базой данных:", e)
            raise

    def load_all_data(self, group_ids):
        """
        Загружает все необходимые данные из базы данных и заполняет локальный словарь.

        :param group_ids: Список двухзначных group_id из Excel.
        """
        cursor = self.conn.cursor()

        # Создаем строку с параметрами для SQL IN
        placeholders = ','.join(['?'] * len(group_ids))
        sql_query = f"""
            SELECT 
                MTR.[код СКМТР], 
                MTR.[Наименование], 
                MTR.[Маркировка], 
                MTR.[Регламенты (ГОСТ/ТУ)], 
                MTR.[Параметры], 
                MTR.[Базисная Единица измерения], 
                OKPD_2.[OKPD2_NAME],
                GOST.[GOST#Gost_title],
                GOST.[GOST#Gost_status],
                GOST.[GOST#Gost_data_start],
                GOST.[GOST#Gost_data_end]
            FROM (MTR
            LEFT JOIN OKPD_2 ON MTR.ОКПД2 = OKPD_2.OKPD2)
            LEFT JOIN GOST ON MTR.[Регламенты (ГОСТ/ТУ)] = GOST.[GOST#Gost_code]
            WHERE LEFT(MTR.ОКПД2, 2) IN ({placeholders})
        """

        try:
            print("Загрузка данных из базы данных...")
            cursor.execute(sql_query, tuple(group_ids))
            rows = cursor.fetchall()
            print(f"Загружено {len(rows)} записей из базы данных.")

            for row in rows:
                # row содержит:
                # (код СКМТР, Наименование, Маркировка, Регламенты (ГОСТ/ТУ),
                # Параметры, Базисная Единица измерения, OKPD2_NAME,
                # Gost_title, Gost_status, Gost_data_start, Gost_data_end)
                id_value = str(row[0]).strip()  # Код СКМТР должен быть строкой
                if not id_value:
                    continue  # Пропускаем пустые ID

                # Пропускаем код СКМТР, так как он используется как ключ
                data_to_store = list(row[1:])  # Срезаем первый элемент

                # Обрабатываем возможные пустые значения и типы данных
                processed_data = []
                for value in data_to_store:
                    if isinstance(value, str):
                        processed_data.append(value.strip())
                    elif isinstance(value, (pyodbc.Timestamp, pyodbc.Date)):
                        processed_data.append(value.strftime('%Y-%m-%d'))
                    else:
                        processed_data.append(value if value is not None else "")

                # Записываем данные в словарь
                self.data_mapping[id_value] = processed_data

        except pyodbc.Error as e:
            print("Ошибка при выполнении SQL-запроса:", e)
            raise
        finally:
            cursor.close()

    def write_data_to_sheet(self, sheet, id_column='A', start_row=2):
        """
        Записывает данные в соответствующие ячейки листа Excel.

        :param sheet: Объект листа Excel.
        :param id_column: Буква столбца, где находятся ID.
        :param start_row: Номер строки, с которой начинаются данные.
        """
        # Заголовки, начиная со второго столбца
        headers = [
            "Наименование", "Маркировка", "Регламенты (ГОСТ/ТУ)",
            "Параметры", "Базисная Единица измерения", "OKPD2_NAME",
            "ГОСТ Название", "ГОСТ Статус", "ГОСТ Дата начала", "ГОСТ Дата окончания"
        ]

        # Проверяем, есть ли уже заголовки (допустим, что заголовки начинаются со второго столбца)
        existing_headers = [cell.value for cell in sheet[1][1:11]]  # Проверяем первые 10 столбцов после ID

        if not any(existing_headers):
            # Добавляем заголовки, если они отсутствуют
            for col_num, header in enumerate(headers, start=2):
                sheet.cell(row=1, column=col_num, value=header)
            print(f"Заголовки добавлены на лист '{sheet.title}'.")
        else:
            print(f"Заголовки уже существуют на листе '{sheet.title}'. Пропускаем добавление заголовков.")

        # Проходимся по всем строкам с ID
        for row in range(start_row, sheet.max_row + 1):
            id_cell = sheet[f"{id_column}{row}"]
            id_value = str(id_cell.value).strip() if id_cell.value is not None else ""
            if not id_value:
                continue  # Пропускаем пустые ячейки

            # Получаем данные из словаря
            data = self.data_mapping.get(id_value, [""] * 10)

            # Записываем данные, начиная со второго столбца
            for col_offset, value in enumerate(data, start=2):
                sheet.cell(row=row, column=col_offset, value=value)

    def process_group_ids(self):
        """
        Основная функция, которая обрабатывает все group_id и записывает объединенные данные в Excel.
        """
        try:
            # Получаем список всех group_id из Excel
            group_ids = self.excel_manager.get_group_ids_from_sheet()
            print(f"Получено {len(group_ids)} group_id из Excel.")

            # Загрузка всех данных из базы данных за один запрос
            self.load_all_data(group_ids)

            # Загружаем Excel-файл
            workbook = openpyxl.load_workbook(self.excel_manager.filename)

            for group_id in group_ids:
                sheet_name = str(group_id)
                if sheet_name not in workbook.sheetnames:
                    print(f"Лист '{sheet_name}' не найден в Excel-файле. Пропускаем.")
                    continue

                sheet = workbook[sheet_name]
                print(f"Обрабатываем лист: {sheet_name}")

                # Записываем данные в лист
                self.write_data_to_sheet(sheet)

                # Сохраняем файл после обработки каждого листа
                workbook.save(self.excel_manager.filename)
                print(f"Лист '{sheet_name}' успешно обновлён и сохранён.")

        except Exception as e:
            print("Произошла ошибка при обработке данных:", e)
        finally:
            self.conn.close()
            print("Соединение с базой данных закрыто.")


# Пример использования
if __name__ == "__main__":
    db_path = r"C:\Hackaton_October\neclassic\main_data.accdb"
    excel_filename = 'main_data.xlsx'

    # Создаем объект класса ExcelGatherDataOptimized
    db_handler = ExcelGatherDataOptimized(db_path, excel_filename)

    # Запускаем процесс обработки данных
    db_handler.process_group_ids()
