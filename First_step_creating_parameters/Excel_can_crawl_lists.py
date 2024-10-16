import openpyxl


class ExcelSheetManager:
    """
    Класс для работы с Excel-файлами, включая получение данных из листа и создание новых листов.
    """

    def __init__(self, filename):
        """
        Инициализирует объект класса ExcelSheetManager с именем файла Excel.

        :param filename: Имя файла Excel для работы.
        """
        self.filename = filename

    def get_group_ids_from_sheet(self, sheet_name='Лист1'):
        """
        Получает group_id из первого столбца листа Excel, игнорируя пустые или None значения.

        :param sheet_name: Имя листа, из которого будут извлекаться данные. По умолчанию 'Лист1'.
        :return: Список значений group_id из первого столбца.
        """
        # Открываем Excel файл
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[sheet_name]

        # Инициализируем список для хранения group_id
        group_ids = []

        # Читаем значения из первой колонки (колонка 'A'), исключая None
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:  # Проверяем, что значение не None
                group_ids.append(row[0])

        return group_ids

    def create_sheet_if_not_exists(self, group_id):
        """
        Проверяет, существует ли лист с именем group_id, если нет - создает новый пустой лист.

        :param group_id: Название листа (group_id), который нужно создать, если он отсутствует.
        """
        # Открываем Excel файл
        workbook = openpyxl.load_workbook(self.filename)

        # Проверяем, существует ли лист с именем group_id
        if str(group_id) not in workbook.sheetnames:
            # Создаем новый лист с названием group_id
            workbook.create_sheet(title=str(group_id))
            print(f'Лист с именем {group_id} создан.')

        # Сохраняем изменения в файле
        workbook.save(self.filename)

    def write_ids_to_sheet(self, group_id, ids):
        """
        Записывает айдишники в указанный лист Excel. Добавляет заголовок 'ID', если его нет.

        :param group_id: Имя листа, в который нужно записать данные.
        :param ids: Список айдишников для записи.
        """
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[str(group_id)]

        # Проверяем наличие заголовка в первой строке и первом столбце
        if sheet.cell(row=1, column=1).value != "ID":
            sheet.cell(row=1, column=1, value="ID")

        # Записываем айдишники, начиная со второй строки, так как первая строка — это заголовок
        for idx, item in enumerate(ids, start=2):
            sheet.cell(row=idx, column=1, value=item)

        workbook.save(self.filename)
        print(f'{len(ids)} айдишников записано в лист {group_id}.')

    def process_excel_file(self):
        """
        Основная функция, которая последовательно вызывает другие функции:
        1. Получает список group_id из файла Excel.
        2. Проверяет и создает листы, если они отсутствуют.
        """
        # Получаем список group_id из первого листа
        group_ids = self.get_group_ids_from_sheet()

        # Создаем листы по каждому group_id, если они не существуют
        for group_id in group_ids:
            self.create_sheet_if_not_exists(group_id)

        print('Процесс обработки файла завершен.')


# Пример использования
if __name__ == "__main__":
    # Укажите имя вашего файла Excel
    filename = 'main_data_with_analysis.xlsx'

    # Создаем объект класса
    manager = ExcelSheetManager(filename)

    # Запускаем процесс обработки Excel-файла
    manager.process_excel_file()
