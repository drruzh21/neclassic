import openpyxl
from First_step_creating_parameters.Excel_can_crawl_lists import ExcelSheetManager
from First_step_creating_parameters.First_GPT_Agent import GptService


class GptDataProcessor:
    """
    Класс для обработки данных из столбца H в Excel и передачи их в GPT сервис.
    Также записывает результаты GPT в столбец I, и сохраняет файл после каждой вставки.
    Обрабатывает только первые 50 строк каждого листа.
    """

    def __init__(self, excel_filename, gpt_service=None):
        """
        Инициализация объекта класса.

        :param excel_filename: Имя файла Excel.
        :param gpt_service: Экземпляр класса GptService для взаимодействия с GPT. Если None, создаётся новый.
        """
        self.excel_manager = ExcelSheetManager(excel_filename)
        self.excel_filename = excel_filename
        self.gpt_service = gpt_service if gpt_service else GptService()

    def process_data(self):
        """
        Основная функция, которая обрабатывает данные из столбца H и передает их в GPT.
        Затем записывает результаты в столбец I и сохраняет Excel-файл после каждой вставки.
        Обрабатывает только первые 50 строк каждого листа.
        """
        try:
            # Загружаем Excel-файл с возможностью записи
            workbook = openpyxl.load_workbook(self.excel_filename)

            # Получаем список всех group_id из Excel (предполагается, что они соответствуют названиям листов)
            group_ids = self.excel_manager.get_group_ids_from_sheet()
            print(f"Получено {len(group_ids)} group_id из Excel.")

            for group_id in group_ids:
                sheet_name = str(group_id)
                if sheet_name not in workbook.sheetnames:
                    print(f"Лист '{sheet_name}' не найден в Excel-файле. Пропускаем.")
                    continue

                sheet = workbook[sheet_name]
                print(f"Обрабатываем лист: {sheet_name}")

                # Получаем заголовок столбца H (8) и I (9)
                header_H = sheet.cell(row=1, column=8)
                if header_H.value != "Собранные данные":
                    # Добавляем заголовок, если его нет
                    header_H.value = "Собранные данные"
                    print(f"Заголовок столбца H добавлен на листе '{sheet_name}'.")

                header_I = sheet.cell(row=1, column=9)
                if header_I.value != "GPT Result":
                    # Добавляем заголовок, если его нет
                    header_I.value = "GPT Result"
                    print(f"Заголовок столбца I добавлен на листе '{sheet_name}'.")

                # Счётчик обработанных строк
                processed_rows = 0

                # Проходимся по всем строкам, начиная со второй, до достижения 50
                for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
                    if processed_rows >= 50:
                        print(f"Достигнуто ограничение в 50 строк для листа '{sheet_name}'. Прекращаем обработку.")
                        break

                    data_cell = row[0]
                    row_number = data_cell.row
                    id_cell = sheet.cell(row=row_number, column=1)  # Столбец A для ID

                    if id_cell.value is None:
                        continue  # Пропускаем строки без ID

                    data_value = data_cell.value
                    if data_value is None or not str(data_value).strip():
                        # Если данных нет, создаём строку с "Неизвестно"
                        formatted_data = (
                            "Наименование: Неизвестно, Маркировка: Неизвестно, "
                            "Регламенты (ГОСТ/ТУ): Неизвестно, Параметры: Неизвестно, "
                            "OKPD2_NAME: Неизвестно, ГОСТ Название: Неизвестно"
                        )
                    else:
                        formatted_data = str(data_value).strip()

                    # Передаём formatted_data в GPT
                    try:
                        result = self.gpt_service.gpt(
                            input_data=formatted_data,
                            system_prompt_file_name="first_prompt_system",
                            basic_prompt_file_name="basic_prompt"
                        )
                    except Exception as gpt_error:
                        print(f"Ошибка при обращении к GPT для ID {id_cell.value}: {gpt_error}")
                        parsed_result = "Ошибка GPT"
                    else:
                        # Получаем parsed_result, если он есть
                        parsed_result = result.get('parsed_result', "Неизвестно")
                        if not parsed_result:
                            parsed_result = "Неизвестно"

                    # Записываем parsed_result в столбец I (9)
                    sheet.cell(row=row_number, column=9, value=parsed_result)

                    # Сохраняем файл после каждой вставки
                    try:
                        workbook.save(self.excel_filename)
                        print(f"Записан результат GPT для ID {id_cell.value} в строку {row_number}. Файл сохранён.")
                    except Exception as save_error:
                        print(f"Ошибка при сохранении файла для ID {id_cell.value}, строка {row_number}: {save_error}")
                        # Решите, хотите ли вы продолжать или прервать выполнение
                        # Например, можно продолжить:
                        continue

                    # Выводим результат в консоль (опционально)
                    print(
                        f"ID: {id_cell.value}, GPT Result: {parsed_result}, Token Usage: {result.get('token_usage', 'N/A')}"
                    )

                    # Увеличиваем счётчик обработанных строк
                    processed_rows += 1

            print("Все листы успешно обработаны и сохранены.")

        except Exception as e:
            print("Произошла ошибка при обработке данных:", e)


# Пример использования
if __name__ == "__main__":
    excel_filename = 'main_data.xlsx'

    # Создаем экземпляр GptService
    gpt_service = GptService()

    # Создаем экземпляр GptDataProcessor
    processor = GptDataProcessor(excel_filename, gpt_service)

    # Запускаем процесс обработки данных
    processor.process_data()
