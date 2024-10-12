import openpyxl
from First_step_creating_parameters.Excel_can_crawl_lists import ExcelSheetManager


class ExcelDataConcatenator:
    """
    Класс для агрегации данных из строк Excel-листа и записи их в столбец H.
    """

    def __init__(self, excel_filename):
        """
        Инициализация объекта класса.

        :param excel_filename: Имя файла Excel.
        """
        self.excel_manager = ExcelSheetManager(excel_filename)

    def concatenate_data(self):
        """
        Основная функция, которая собирает данные из каждой строки и записывает их в столбец H.
        """
        try:
            # Загружаем Excel-файл
            workbook = openpyxl.load_workbook(self.excel_manager.filename)

            # Получаем список всех group_id из Excel (предполагается, что они соответствуют названиям листов)
            group_ids = self.excel_manager.get_group_ids_from_sheet()
            print(f"Получено {len(group_ids)} group_id из Excel.")

            for group_id in group_ids:
                sheet_name = str(group_id)
                if sheet_name not in workbook.sheetnames:
                    print(f"Лист '{sheet_name}' не найден в Excel-файле. Пропускаем.")
                    continue

                sheet = workbook[sheet_name]
                print(f"Обрабатываем лист: {sheet_name}")

                # Получаем заголовки из первой строки
                headers = [cell.value for cell in sheet[1]]

                # Определяем диапазон столбцов A-G (индексы 0-6)
                data_columns = headers[:7]  # Предполагается, что первые 7 столбцов содержат необходимые данные

                # Проверяем наличие заголовков
                if not any(data_columns):
                    print(f"Заголовки отсутствуют на листе '{sheet_name}'. Пропускаем.")
                    continue

                # Добавляем заголовок в столбец H, если его нет
                if sheet.cell(row=1, column=8).value is None:
                    sheet.cell(row=1, column=8, value="Собранные данные")
                    print(f"Заголовок добавлен в столбец H на листе '{sheet_name}'.")

                # Обрабатываем каждую строку, начиная со второй
                for row in sheet.iter_rows(min_row=2, max_col=7):
                    id_cell = row[0]  # Предполагается, что ID находится в столбце A
                    if id_cell.value is None:
                        continue  # Пропускаем строки без ID

                    # Собираем данные из столбцов A-G
                    data_values = [cell.value for cell in row]

                    # Заменяем отсутствующие значения на "Неизвестно"
                    processed_values = [str(value).strip() if value is not None else "Неизвестно" for value in data_values]

                    # Создаем строку с данными, добавляя заголовки
                    data_string = ", ".join([f"{headers[i]}: {processed_values[i]}" for i in range(len(processed_values))])

                    # Записываем строку в столбец H
                    sheet.cell(row=id_cell.row, column=8, value=data_string)

                # Сохраняем файл после обработки каждого листа
                workbook.save(self.excel_manager.filename)
                print(f"Лист '{sheet_name}' успешно обновлён и сохранён.")

        except Exception as e:
            print("Произошла ошибка при агрегации данных:", e)

# Пример использования
if __name__ == "__main__":
    excel_filename = 'main_data.xlsx'

    # Создаем объект класса ExcelDataConcatenator
    aggregator = ExcelDataConcatenator(excel_filename)

    # Запускаем процесс агрегации данных
    aggregator.concatenate_data()
