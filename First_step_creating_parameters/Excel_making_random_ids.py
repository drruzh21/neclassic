import openpyxl
import pyodbc
import random

from First_step_creating_parameters.Excel_can_crawl_lists import ExcelSheetManager


class DatabaseHandler:
    """
    Класс для работы с базой данных Access и выбора случайных айдишников по ОКПД2.
    """

    def __init__(self, db_path, excel_filename):
        """
        Инициализирует объект класса DatabaseHandler.

        :param db_path: Путь к базе данных Access.
        :param excel_filename: Имя файла Excel.
        """
        self.db_path = db_path
        self.excel_manager = ExcelSheetManager(excel_filename)

    def fetch_ids_by_okpd2(self, group_id):
        """
        Извлекает айдишники из базы данных, у которых первые две цифры ОКПД2 совпадают с group_id.

        :param group_id: Первые две цифры ОКПД2 для фильтрации.
        :return: Список айдишников (значения из поля 'код СКМТР').
        """
        conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                r'DBQ=' + self.db_path + ';'
        )
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # SQL-запрос для извлечения айдишников по первым двум цифрам ОКПД2
        query = "SELECT [код СКМТР] FROM MTR WHERE LEFT([ОКПД2], 2) = ?"

        # Передаем параметр group_id
        cursor.execute(query, (group_id,))

        # Получаем все подходящие строки
        ids = [row[0] for row in cursor.fetchall()]

        # Закрываем соединение
        cursor.close()
        conn.close()

        return ids


    def select_random_ids(self, ids, count):
        """
        Выбирает случайное количество айдишников из списка.

        :param ids: Список всех доступных айдишников.
        :param count: Количество айдишников для выбора.
        :return: Список случайных айдишников.
        """
        return random.sample(ids, min(count, len(ids)))

    def process_group_ids(self):
        """
        Основной метод, который обрабатывает все group_id, извлекает данные и записывает их в Excel.
        """
        # Получаем список всех group_id из Excel
        group_ids = self.excel_manager.get_group_ids_from_sheet()

        for group_id in group_ids:
            # Получаем количество, которое нужно использовать, и общее количество (например, count_to_use и count)
            ids = self.fetch_ids_by_okpd2(group_id)

            # Загружаем количество строк для выборки (например, count_to_use)
            workbook = openpyxl.load_workbook(self.excel_manager.filename)
            sheet = workbook['Лист1']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == group_id:
                    count_to_use = row[2]  # Столбец count_to_use
                    break

            # Выбираем случайные айдишники
            selected_ids = self.select_random_ids(ids, count_to_use)

            # Создаем лист, если его нет, и записываем данные
            self.excel_manager.create_sheet_if_not_exists(group_id)
            self.excel_manager.write_ids_to_sheet(group_id, selected_ids)


# Пример использования
if __name__ == "__main__":
    db_path = r"C:\Hackaton_October\neclassic\main_data.accdb"
    excel_filename = 'main_data.xlsx'

    # Создаем объект класса DatabaseHandler
    db_handler = DatabaseHandler(db_path, excel_filename)

    # Запускаем процесс
    db_handler.process_group_ids()
