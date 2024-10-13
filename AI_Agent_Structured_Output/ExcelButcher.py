import openpyxl
from typing import List, Dict
from math import ceil

from AI_Agent_Structured_Output.MainAgent import AiAgentStructuredOutput
from AI_Agent_Structured_Output.Structures.WarCrimes import ProductCriteria14, ProductCriteria27, ProductCriteria28, \
    ProductCriteria25, ProductCriteria30, ProductCriteria26, ProductCriteria22, ProductCriteria29, ProductCriteria20, \
    ProductCriteria31, ProductCriteria23, ProductCriteria32, ProductCriteria58, ProductCriteria17, ProductCriteria15, \
    ProductCriteria10, ProductCriteria24, ProductCriteria13, ProductCriteria16, ProductCriteria11, ProductCriteria21, \
    ProductCriteria19, ProductCriteria01
from First_step_creating_parameters.Excel_can_crawl_lists import ExcelSheetManager

# Определяем словарь соответствия названий листов и классов критериев
product_criteria_dict = {
    "14": ProductCriteria14(),
    "27": ProductCriteria27(),
    "28": ProductCriteria28(),
    "25": ProductCriteria25(),
    "30": ProductCriteria30(),
    "26": ProductCriteria26(),
    "22": ProductCriteria22(),
    "29": ProductCriteria29(),
    "20": ProductCriteria20(),
    "31": ProductCriteria31(),
    "32": ProductCriteria32(),
    "23": ProductCriteria23(),
    "58": ProductCriteria58(),
    "17": ProductCriteria17(),
    "15": ProductCriteria15(),
    "10": ProductCriteria10(),
    "24": ProductCriteria24(),
    "13": ProductCriteria13(),
    "16": ProductCriteria16(),
    "11": ProductCriteria11(),
    "21": ProductCriteria21(),
    "19": ProductCriteria19(),
    "01": ProductCriteria01()
}


class ExcelBatchProcessor:
    """
    Класс для обработки Excel-файлов, формирования батчей данных и отправки их на анализ.
    Использует композицию с классом ExcelSheetManager для управления листами Excel-файла.
    """

    def __init__(self, filename: str, batch_size: int = 10):
        """
        Инициализирует ExcelBatchProcessor.

        :param filename: Имя Excel-файла для обработки.
        :param batch_size: Количество элементов в одном батче.
        """
        self.filename = filename
        self.batch_size = batch_size
        self.manager = ExcelSheetManager(filename)
        try:
            self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
        except FileNotFoundError:
            print(f"Файл {self.filename} не найден.")
            exit(1)
        except Exception as e:
            print(f"Ошибка при загрузке файла {self.filename}: {e}")
            exit(1)
        self.agent = AiAgentStructuredOutput()

    def collect_data_per_sheet(self, sheet_name: str) -> List[List[str]]:
        """
        Считывает данные из указанного листа Excel-файла, извлекает первые 7 значений каждой строки
        и добавляет название листа в начало массива.

        :param sheet_name: Название листа Excel для обработки.
        :return: Список массивов с данными для отправки в AI агент.
        """
        collected_data = []
        sheet = self.workbook[sheet_name]
        print(f'Обработка листа: {sheet_name}')

        # Предполагается, что первая строка - заголовок
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            if all(cell is None for cell in row):
                # Пропускаем полностью пустые строки
                continue

            # Преобразуем значения ячеек в строки, заменяя None на пустую строку
            row_data = [str(cell) if cell is not None else "" for cell in row]

            # Добавляем название листа в начало массива
            data_entry = [sheet_name] + row_data
            collected_data.append(data_entry)

        print(f'Всего собрано записей из листа {sheet_name}: {len(collected_data)}')
        return collected_data

    def create_batches(self, data: List[List[str]]) -> List[List[List[str]]]:
        """
        Формирует батчи из собранных данных.

        :param data: Список собранных данных.
        :return: Список батчей, каждый батч - список из batch_size элементов.
        """
        total_batches = ceil(len(data) / self.batch_size)
        batches = []

        for i in range(total_batches):
            batch = data[i * self.batch_size: (i + 1) * self.batch_size]
            batches.append(batch)
            print(f'  Сформирован батч {i + 1} с {len(batch)} элементами.')

        return batches

    def collect_batches(self) -> Dict[str, List[List[List[str]]]]:
        """
        Собирает все батчи из всех листов Excel-файла.

        :return: Словарь с именами листов и их соответствующими списками батчей.
        """
        all_batches = {}

        for sheet_name in self.workbook.sheetnames:
            if sheet_name not in product_criteria_dict:
                print(f'Нет критериев для листа {sheet_name}, пропуск.')
                continue

            criteria = product_criteria_dict[sheet_name]
            data = self.collect_data_per_sheet(sheet_name)

            if not data:
                print(f'Нет данных для обработки в листе {sheet_name}.')
                continue

            batches = self.create_batches(data)
            all_batches[sheet_name] = batches

        return all_batches

    def send_batches_to_agent(self, all_batches: Dict[str, List[List[List[str]]]]):
        """
        Отправляет все батчи в AiAgentStructuredOutput для обработки.

        :param all_batches: Словарь с именами листов и их соответствующими списками батчей.
        """
        for sheet_name, batches in all_batches.items():
            criteria = product_criteria_dict[sheet_name]
            for idx, batch in enumerate(batches, start=1):
                # Удаляем первый элемент (название листа) из каждого массива данных в батче
                modified_batch = [item[1:] for item in batch]
                print(f'Отправка батча {idx} из листа {sheet_name} в агента...')
                self.agent.structured_output_agent(modified_batch, criteria, sheet_name)
                print(f'Батч {idx} из листа {sheet_name} обработан.\n')

    def send_single_batch_to_agent(self, sheet_name: str, batch_index: int):
        """
        Отправляет один конкретный батч в AiAgentStructuredOutput для обработки.

        :param sheet_name: Название листа Excel.
        :param batch_index: Номер батча (начинается с 1).
        """
        if sheet_name not in product_criteria_dict:
            print(f'Нет критериев для листа {sheet_name}, невозможно обработать батч.')
            return

        batches = self.collect_batches().get(sheet_name, [])

        if not batches:
            print(f'Нет батчей для листа {sheet_name}.')
            return

        if batch_index < 1 or batch_index > len(batches):
            print(f'Неверный номер батча {batch_index} для листа {sheet_name}.')
            return

        criteria = product_criteria_dict[sheet_name]
        batch = batches[batch_index - 1]

        # Удаляем первый элемент (название листа) из каждого массива данных в батче
        modified_batch = [item[1:] for item in batch]

        print(f'Отправка батча {batch_index} из листа {sheet_name} в агента...')
        self.agent.structured_output_agent(modified_batch, criteria)
        print(f'Батч {batch_index} из листа {sheet_name} обработан.\n')


if __name__ == "__main__":
    # Укажите путь к вашему файлу Excel
    filename = 'C:/Hackaton_October/neclassic/First_step_creating_parameters/main_data_with_analysis.xlsx'

    # Создаем объект класса ExcelBatchProcessor с размером батча 10
    processor = ExcelBatchProcessor(filename=filename, batch_size=2)

    # Собираем все батчи из Excel
    all_batches = processor.collect_batches()

    # Выводим информацию о сформированных батчах для каждого листа
    for sheet_name, batches in all_batches.items():
        print(f'\nЛист: {sheet_name}, Всего батчей: {len(batches)}')
        for batch_index, batch in enumerate(batches, start=1):
            print(f'  Батч {batch_index}:')
            for item in batch:
                print(f'    {item}')
            print('\n')  # Разделение между батчами для читаемости

    # Предлагаем пользователю выбрать, что делать дальше
    while True:
        action = input("Вы хотите отправить батчи в агента? (все - 'a', один - 'o', выйти - 'q'): ").strip().lower()
        if action == 'a':
            processor.send_batches_to_agent(all_batches)
            print('Все батчи отправлены в агента.')
            break
        elif action == 'o':
            # Запрашиваем название листа
            sheet = input("Введите название листа для обработки: ").strip()
            if sheet not in all_batches:
                print(f'Лист "{sheet}" не найден или не содержит батчей.')
                continue
            # Запрашиваем номер батча
            try:
                batch_num = int(input(f"Введите номер батча для листа {sheet} (1-{len(all_batches[sheet])}): ").strip())
            except ValueError:
                print("Пожалуйста, введите корректный номер батча.")
                continue
            # Отправляем выбранный батч
            processor.send_single_batch_to_agent(sheet, batch_num)
        elif action == 'q':
            print('Выход из программы.')
            break
        else:
            print(
                "Неверный ввод. Пожалуйста, выберите 'a' для отправки всех батчей, 'o' для отправки одного батча или 'q' для выхода.")
