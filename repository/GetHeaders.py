import openpyxl
import json


def get_headers_from_excel():
    file_path = 'main_data_with_analysis.xlsx'

    workbook = openpyxl.load_workbook(file_path)

    j2_values = []
    table_headers = []
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        j2_value = ws['J2'].value
        if j2_value is not None:
            j2_values.append((sheet, j2_value))
        else:
            j2_values.append((sheet, 'Ячейка J2 пуста'))

    for sheet, value in j2_values:
        if '<top_criteria>' in value and '</top_criteria>' in value:
            start = value.find('<top_criteria>') + len('<top_criteria>')
            end = value.find('</top_criteria>')
            json_str = value[start:end].strip()
            res = [sheet, 'ID', 'Наименование', 'Маркировка', 'Регламенты (ГОСТ/ТУ)', 'Параметры', 'OKPD2_NAME',
                   'ГОСТ Название']
            try:
                json_data = json.loads(json_str)
                top_criteria_values = json_data.get("top_criteria", [])
                if len(top_criteria_values) > 0:
                    for item in top_criteria_values:
                        res.append(item)
                    table_headers.append(res)
            except json.JSONDecodeError as e:
                print("Ошибка при парсинге JSON:", e)
    return table_headers
